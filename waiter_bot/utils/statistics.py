import pandas as pl
from datetime import datetime

from loader import bot
import os
import openpyxl


async def write_excel(queryset, chat_id):
    """Write data to Excel file"""
    if not os.path.exists("stats_files"):
        os.makedirs("stats_files")

    file_path = (
        f"stats_files/statistics_{datetime.strftime(datetime.now(), '%Y-%m-%d_%H-%M-%S')}.xlsx"
    )
    pl.DataFrame(
        {
            "№": [i + 1 for i in range(len(queryset))],
            "Buyurtma vaqti": [datetime.strftime(i["ordered_at"], "%Y-%m-%d %H:%M:%S") for i in queryset],
            "Stol": [i["table_name"] for i in queryset],
            "Mahsulotlar": [i["products"] for i in queryset],
            "Ofitsiant": [i["waiter"] for i in queryset],
            "Mahsulotlar narxi Jami": [i["total_products_price"] for i in queryset],
            "Ofitsiant xizmati": [i["service_fee"] for i in queryset],
            "Jami": [i["total_price"] for i in queryset],
        }
    ).write_excel(file_path, column_totals=True, autofit=True)
    await bot.send_document(chat_id=chat_id, document=open(file_path, "rb"))


async def write_general_statistics_excel(chat_id, best_selling_product, general_statistics, queryset):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.merge_cells('A1:C1')
    sheet['A1'] = 'Eng ko\'p sotilgan mahsulot'
    sheet['A1'].alignment = openpyxl.styles.Alignment(horizontal='center')
    sheet.append(["Nomi", "Jami sotilgan", "Umumiy narxi"])
    for i in ['A2', 'B2', 'C2']:
        sheet[i].alignment = openpyxl.styles.Alignment(horizontal='center')
    sheet.append([best_selling_product['name'], int(best_selling_product['total_weight']),
                  best_selling_product['total_price']])
    sheet.append(["", "", ""])
    sheet.merge_cells('A5:D5')
    sheet['A5'] = 'Umumiy statistika'
    sheet['A5'].alignment = openpyxl.styles.Alignment(horizontal='center')
    sheet.append(["Buyurtmalar soni", "Jami daromadi", "Xizmatlar summasi", "Jami summa"])
    for i in ['A6', 'B6', 'C6', 'D6']:
        sheet[i].alignment = openpyxl.styles.Alignment(horizontal='center')
    sheet.append([general_statistics['total_orders'], general_statistics['total_price'],
                  general_statistics['service_fee'], general_statistics['total']])
    sheet.merge_cells('A9:G9')
    sheet['A9'] = "Barcha sotuvlar"
    sheet['A9'].alignment = openpyxl.styles.Alignment(horizontal='center')
    sheet.append(["Buyurtma vaqti", "Stol / Xona", "Mahsulotlar", "Ofitsiant", "Mahsulotlar narxi Jami",
                  "Ofitsiant xizmati", "Jami"])
    for i in ['A10', 'B10', 'C10', 'D10', 'E10', 'F10', 'G10']:
        sheet[i].alignment = openpyxl.styles.Alignment(horizontal='center')
    for i in queryset:
        sheet.append([datetime.strftime(i["ordered_at"], "%Y-%m-%d %H:%M:%S"), i["table_name"], i["products"],
                      i["waiter"], i["total_products_price"], i["service_fee"], i["total_price"]])
    workbook.save('output_excel_file.xlsx')
    return open('output_excel_file.xlsx', "rb")


async def write_by_waiter_statistics_excel(chat_id, queryset):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Ofitsiant", "Buyurtmalar soni", "Jami daromadi", "Xizmatlar summasi", "Jami summa"])
    for i in ['A1', 'B1', 'C1', 'D1', 'E1']:
        sheet[i].alignment = openpyxl.styles.Alignment(horizontal='center')
    sheet.append([queryset['waiter'], queryset['total_orders'], queryset['total_price'], queryset['service_fee'],
                  queryset['total']])

    # Save the workbook to a file
    workbook.save('output_excel_file.xlsx')
    return open('output_excel_file.xlsx', "rb")


async def write_by_product_statistics_excel(chat_id, queryset):
    workbook = openpyxl.Workbook()
    msg = "Mahsulotlar statistikasi\n\n"
    sheet = workbook.active
    sheet.append(["Mahsulot", "Buyurtmalar soni", "Jami sotuv"])
    for i in ['A1', 'B1', 'C1']:
        sheet[i].alignment = openpyxl.styles.Alignment(horizontal='center')
    for q in queryset:
        sheet.append([q['product_name'], q['total_weight'], q['total_price']])
    workbook.save('output_excel_file.xlsx')
    return open('output_excel_file.xlsx', "rb")
